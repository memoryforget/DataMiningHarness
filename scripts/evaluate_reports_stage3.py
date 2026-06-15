#!/usr/bin/env python3

import argparse
import csv
import json
import re
import sys
import time
import urllib.error
import urllib.request
from copy import deepcopy
from pathlib import Path
from typing import Any


TEXT_SUFFIXES = {".txt", ".md", ".log", ".rst"}
CODE_SUFFIXES = {".py", ".r", ".sql", ".sh", ".js", ".ts", ".ipynb"}
TABLE_SUFFIXES = {".csv", ".tsv"}
JSON_SUFFIXES = {".json", ".geojson"}
JSONL_SUFFIXES = {".jsonl", ".ndjson"}
MAX_REPORT_CHARS = 60000
MAX_PACKET_CHARS = 70000
MAX_CELL_CHARS = 180


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--task-evaluation-json", required=True)
    parser.add_argument("--candidate-dir", required=True)
    parser.add_argument("--output-json", required=True)
    parser.add_argument("--judge-api-url", required=True)
    parser.add_argument("--judge-api-key", required=True)
    parser.add_argument("--judge-model", required=True)
    parser.add_argument("--temperature", type=float, default=0.0)
    parser.add_argument("--max-output-tokens", type=int, default=4000)
    return parser.parse_args()


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def read_text_best_effort(path: Path, max_chars: int | None = None) -> str:
    text = path.read_text(encoding="utf-8", errors="replace")
    if max_chars is not None and len(text) > max_chars:
        return text[:max_chars] + "\n...[truncated]"
    return text


def truncate(value: Any, max_chars: int = MAX_CELL_CHARS) -> Any:
    if isinstance(value, str) and len(value) > max_chars:
        return value[:max_chars] + "..."
    return value


def extract_keywords(*texts: str) -> list[str]:
    stop = {
        "about",
        "after",
        "also",
        "and",
        "are",
        "artifact",
        "because",
        "been",
        "check",
        "data",
        "does",
        "field",
        "from",
        "has",
        "have",
        "into",
        "not",
        "output",
        "report",
        "required",
        "rubric",
        "that",
        "the",
        "this",
        "using",
        "with",
    }
    counts: dict[str, int] = {}
    for text in texts:
        for token in re.findall(r"[A-Za-z_][A-Za-z0-9_]{2,}", text.lower()):
            if token in stop:
                continue
            counts[token] = counts.get(token, 0) + 1
    return [k for k, _ in sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:40]]


def keyword_hits_from_lines(lines: list[str], keywords: list[str], limit: int = 20) -> list[dict[str, Any]]:
    hits: list[dict[str, Any]] = []
    lowered_keywords = [k.lower() for k in keywords]
    for idx, line in enumerate(lines, start=1):
        lowered = line.lower()
        matched = [k for k in lowered_keywords if k in lowered]
        if not matched:
            continue
        hits.append(
            {
                "line": idx,
                "matched_keywords": matched[:8],
                "text": truncate(line.strip(), 500),
            }
        )
        if len(hits) >= limit:
            break
    return hits


def summarize_table(path: Path, delimiter: str, keywords: list[str]) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    raw_lines: list[str] = []
    row_count = 0
    columns: list[str] = []
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as fh:
        raw_lines = fh.readlines()
    reader = csv.DictReader(raw_lines, delimiter=delimiter)
    columns = list(reader.fieldnames or [])
    for row in reader:
        row_count += 1
        if len(rows) < 12:
            rows.append({str(k): truncate(v) for k, v in row.items()})
    return {
        "kind": "table",
        "columns": columns,
        "row_count": row_count,
        "sample_rows": rows,
        "keyword_hits": keyword_hits_from_lines(raw_lines, keywords),
    }


def json_schema_sample(obj: Any, depth: int = 0) -> Any:
    if depth >= 3:
        return type(obj).__name__
    if isinstance(obj, dict):
        return {str(k): json_schema_sample(v, depth + 1) for k, v in list(obj.items())[:30]}
    if isinstance(obj, list):
        return {
            "type": "list",
            "length": len(obj),
            "sample": [json_schema_sample(v, depth + 1) for v in obj[:3]],
        }
    return type(obj).__name__


def summarize_json(path: Path, keywords: list[str]) -> dict[str, Any]:
    text = read_text_best_effort(path, 250000)
    try:
        obj = json.loads(text)
    except json.JSONDecodeError as exc:
        return {
            "kind": "json",
            "parse_error": str(exc),
            "first_lines": text.splitlines()[:80],
            "keyword_hits": keyword_hits_from_lines(text.splitlines(), keywords),
        }
    return {
        "kind": "json",
        "schema_sample": json_schema_sample(obj),
        "sample": truncate(json.dumps(obj, ensure_ascii=False)[:6000], 6000),
        "keyword_hits": keyword_hits_from_lines(text.splitlines(), keywords),
    }


def summarize_jsonl(path: Path, keywords: list[str]) -> dict[str, Any]:
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    samples: list[Any] = []
    key_union: set[str] = set()
    parse_errors: list[str] = []
    for idx, line in enumerate(lines, start=1):
        if not line.strip():
            continue
        try:
            obj = json.loads(line)
        except json.JSONDecodeError as exc:
            if len(parse_errors) < 5:
                parse_errors.append(f"line {idx}: {exc}")
            continue
        if isinstance(obj, dict):
            key_union.update(str(k) for k in obj.keys())
        if len(samples) < 8:
            samples.append(obj)
    return {
        "kind": "jsonl",
        "line_count": len(lines),
        "key_union": sorted(key_union),
        "sample_records": samples,
        "parse_errors": parse_errors,
        "keyword_hits": keyword_hits_from_lines(lines, keywords),
    }


def summarize_text(path: Path, keywords: list[str], max_lines: int = 100) -> dict[str, Any]:
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    return {
        "kind": "text",
        "line_count": len(lines),
        "first_lines": [truncate(line, 500) for line in lines[:max_lines]],
        "keyword_hits": keyword_hits_from_lines(lines, keywords),
    }


def summarize_code(path: Path, keywords: list[str]) -> dict[str, Any]:
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    definitions = []
    pattern = re.compile(r"^\s*(def|class|function|sub|proc)\s+([A-Za-z_][A-Za-z0-9_]*)")
    for idx, line in enumerate(lines, start=1):
        match = pattern.search(line)
        if match:
            definitions.append({"line": idx, "name": match.group(2), "kind": match.group(1)})
    return {
        "kind": "code",
        "line_count": len(lines),
        "definitions": definitions[:80],
        "first_lines": [truncate(line, 500) for line in lines[:140]],
        "keyword_hits": keyword_hits_from_lines(lines, keywords, limit=30),
    }


def summarize_xlsx(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        return {"kind": "spreadsheet", "summary_error": f"openpyxl unavailable: {exc}"}
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return {"kind": "spreadsheet", "summary_error": str(exc)}
    sheets: list[dict[str, Any]] = []
    for sheet in workbook.worksheets[:5]:
        rows = []
        for row in sheet.iter_rows(max_row=12, values_only=True):
            rows.append([truncate(value) for value in row])
        sheets.append(
            {
                "name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "sample_rows": rows,
            }
        )
    workbook.close()
    return {"kind": "spreadsheet", "sheets": sheets}


def summarize_artifact(path: Path, keywords: list[str]) -> dict[str, Any]:
    stat = path.stat()
    suffix = path.suffix.lower()
    base: dict[str, Any] = {
        "exists": True,
        "size_bytes": stat.st_size,
        "suffix": suffix,
    }
    try:
        if suffix == ".csv":
            base.update(summarize_table(path, ",", keywords))
        elif suffix == ".tsv":
            base.update(summarize_table(path, "\t", keywords))
        elif suffix in JSON_SUFFIXES:
            base.update(summarize_json(path, keywords))
        elif suffix in JSONL_SUFFIXES:
            base.update(summarize_jsonl(path, keywords))
        elif suffix in TEXT_SUFFIXES:
            base.update(summarize_text(path, keywords))
        elif suffix in CODE_SUFFIXES:
            base.update(summarize_code(path, keywords))
        elif suffix in {".xlsx", ".xlsm"}:
            base.update(summarize_xlsx(path))
        else:
            base["kind"] = "binary_or_unsupported"
    except Exception as exc:
        base["summary_error"] = str(exc)
    return base


def safe_artifact_path(candidate_dir: Path, path_text: str) -> Path | None:
    normalized = path_text.strip().replace("\\", "/")
    if not normalized.startswith("artifacts/"):
        return None
    if normalized.startswith("/") or ".." in Path(normalized).parts:
        return None
    resolved = (candidate_dir / normalized).resolve()
    try:
        resolved.relative_to(candidate_dir.resolve())
    except ValueError:
        return None
    return resolved


def collect_uncertain_rubrics(task_eval: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        rubric
        for rubric in task_eval.get("rubric_results", [])
        if isinstance(rubric, dict) and str(rubric.get("verdict", "")).lower() == "uncertain"
    ]


def build_artifact_packet(candidate_dir: Path, uncertain: list[dict[str, Any]]) -> dict[str, Any]:
    artifacts: dict[str, dict[str, Any]] = {}
    for rubric in uncertain:
        texts = [
            str(rubric.get("criteria", "")),
            str(rubric.get("dimension", "")),
            str(rubric.get("reason", "")),
            " ".join(str(x) for x in rubric.get("evidence", [])),
        ]
        keywords = extract_keywords(*texts)
        for item in rubric.get("requested_artifacts", []):
            if isinstance(item, str):
                raw_path = item
                purpose = ""
            elif isinstance(item, dict):
                raw_path = str(item.get("path", ""))
                purpose = str(item.get("purpose", ""))
            else:
                continue
            safe_path = safe_artifact_path(candidate_dir, raw_path)
            if safe_path is None:
                artifacts[raw_path] = {
                    "requested_path": raw_path,
                    "exists": False,
                    "error": "requested artifact path is outside artifacts/ or unsafe",
                    "purpose": purpose,
                }
                continue
            rel_path = safe_path.relative_to(candidate_dir.resolve()).as_posix()
            if rel_path in artifacts:
                artifacts[rel_path].setdefault("purposes", []).append(purpose)
                continue
            if not safe_path.exists():
                artifacts[rel_path] = {
                    "requested_path": rel_path,
                    "exists": False,
                    "error": "artifact does not exist",
                    "purpose": purpose,
                }
                continue
            if safe_path.is_dir():
                entries = sorted(p.relative_to(safe_path).as_posix() for p in safe_path.rglob("*") if p.is_file())[:200]
                artifacts[rel_path] = {
                    "requested_path": rel_path,
                    "exists": True,
                    "kind": "directory",
                    "file_count_sampled": len(entries),
                    "sample_files": entries,
                    "purpose": purpose,
                }
                continue
            summary = summarize_artifact(safe_path, keywords)
            summary["requested_path"] = rel_path
            summary["purpose"] = purpose
            artifacts[rel_path] = summary
    packet = {
        "artifact_root": str(candidate_dir / "artifacts"),
        "artifacts": artifacts,
    }
    packet_text = json.dumps(packet, ensure_ascii=False)
    if len(packet_text) > MAX_PACKET_CHARS:
        packet["truncated"] = True
        packet["artifacts"] = shrink_artifact_packet(artifacts)
    return packet


def shrink_artifact_packet(artifacts: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    shrunk: dict[str, dict[str, Any]] = {}
    for path, summary in artifacts.items():
        compact = dict(summary)
        for key in ("sample_rows", "first_lines", "sample_records"):
            value = compact.get(key)
            if isinstance(value, list):
                compact[key] = value[:5]
        if "sample" in compact and isinstance(compact["sample"], str):
            compact["sample"] = compact["sample"][:1200] + "...[truncated]"
        shrunk[path] = compact
    return shrunk


def post_json(url: str, api_key: str, payload: dict[str, Any]) -> dict[str, Any]:
    body = json.dumps(payload).encode("utf-8")
    last_error: Exception | None = None
    for attempt in range(1, 4):
        req = urllib.request.Request(
            url,
            data=body,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=300) as resp:
                raw = resp.read().decode("utf-8")
            try:
                return json.loads(raw)
            except json.JSONDecodeError:
                last_error = RuntimeError(f"non-JSON response on attempt {attempt}: {raw[:200]!r}")
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            last_error = RuntimeError(f"HTTP {exc.code}: {detail}")
        except urllib.error.URLError as exc:
            last_error = RuntimeError(f"request failed: {exc}")
        if attempt < 3:
            time.sleep(5 * attempt)
    raise RuntimeError(str(last_error) if last_error else "request failed")


def run_judge(prompt: str, api_url: str, api_key: str, model: str, temperature: float, max_output_tokens: int) -> tuple[str, dict[str, Any]]:
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": temperature,
        "max_tokens": max_output_tokens,
    }
    response = post_json(api_url.rstrip("/") + "/chat/completions", api_key, payload)
    choices = response.get("choices")
    if not isinstance(choices, list) or not choices:
        raise ValueError("judge response missing choices")
    content = choices[0].get("message", {}).get("content")
    if not isinstance(content, str) or not content.strip():
        raise ValueError("judge response missing message content")
    return content, response


def parse_judge_json(raw_output: str) -> dict[str, Any]:
    text = raw_output.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}")
        if start == -1 or end == -1 or end <= start:
            raise
        return json.loads(text[start : end + 1])


def build_stage3_prompt(task_eval: dict[str, Any], uncertain: list[dict[str, Any]], report_text: str, packet: dict[str, Any]) -> str:
    schema = {
        "adjudications": [
            {
                "subtask_id": "string",
                "rubric_index": 1,
                "final_passed": True,
                "confidence": "high|medium|low",
                "artifact_evidence": ["specific artifact fields, rows, keys, or lines used"],
                "reason": "short final justification",
            }
        ],
        "overall_notes": ["optional notes"],
    }
    rubric_payload = [
        {
            "subtask_id": r.get("subtask_id"),
            "rubric_index": r.get("rubric_index"),
            "dimension": r.get("dimension"),
            "criteria": r.get("criteria"),
            "stage2_reason": r.get("reason"),
            "stage2_evidence": r.get("evidence", []),
            "requested_artifacts": r.get("requested_artifacts", []),
        }
        for r in uncertain
    ]
    return f"""You are Stage3 of a benchmark judge.

Stage2 could not decide the rubrics below from the report alone. Your job is to resolve only these uncertain rubrics using the report plus summarized contents of the report-cited artifacts.

Rules:
- Use artifacts only to resolve a claim or output that is already present in the report or Stage2 evidence.
- Do not award credit for a completely missing analysis just because an artifact happens to contain related words.
- Mark final_passed true when the report plus artifact summary show the rubric's requirement was satisfied.
- Mark final_passed false when the artifact is missing, irrelevant, contradicts the report, or still does not show the required method/result.
- If a rubric depends on a data-specific edge case, a clear artifact-supported check that found no such records can be sufficient.

Return strict JSON only. Return exactly one adjudication for every uncertain rubric.

Required output schema:
{json.dumps(schema, ensure_ascii=False, indent=2)}

Task metadata:
{json.dumps({k: task_eval.get(k) for k in ["task_id", "query_id", "domain", "agent_facing_prompt"]}, ensure_ascii=False, indent=2)}

Uncertain rubrics:
{json.dumps(rubric_payload, ensure_ascii=False, indent=2)}

Candidate report:
```markdown
{report_text}
```

Artifact packet:
{json.dumps(packet, ensure_ascii=False, indent=2)}
"""


def recompute_summary(task_eval: dict[str, Any]) -> None:
    rubric_results = task_eval.get("rubric_results", [])
    subtask_results = task_eval.get("subtask_results", [])
    for subtask in subtask_results:
        if not isinstance(subtask, dict):
            continue
        subtask_id = subtask.get("subtask_id")
        group = [r for r in rubric_results if isinstance(r, dict) and r.get("subtask_id") == subtask_id]
        passed_rubrics = sum(1 for r in group if r.get("passed"))
        failed_rubrics = [r for r in group if not r.get("passed")]
        subtask["passed"] = bool(group) and passed_rubrics == len(group)
        subtask["passed_rubrics"] = passed_rubrics
        subtask["total_rubrics"] = len(group)
        subtask["failed_rubric_indexes"] = [r.get("rubric_index") for r in failed_rubrics]
        subtask["failed_rubrics"] = failed_rubrics

    total_rubrics = len(rubric_results)
    passed_rubrics = sum(1 for r in rubric_results if isinstance(r, dict) and r.get("passed"))
    total_subtasks = len(subtask_results)
    passed_subtasks = sum(1 for s in subtask_results if isinstance(s, dict) and s.get("passed"))
    task_eval["summary"] = {
        "passed_rubrics": passed_rubrics,
        "total_rubrics": total_rubrics,
        "rubric_coverage": passed_rubrics / total_rubrics if total_rubrics else 0.0,
        "passed_subtasks": passed_subtasks,
        "total_subtasks": total_subtasks,
        "subtask_pass_rate": passed_subtasks / total_subtasks if total_subtasks else 0.0,
        "task_score": passed_subtasks / total_subtasks if total_subtasks else 0.0,
        "task_passed": total_subtasks > 0 and passed_subtasks == total_subtasks,
    }


def apply_adjudications(task_eval: dict[str, Any], judge_result: dict[str, Any]) -> dict[str, Any]:
    result = deepcopy(task_eval)
    adjudications = judge_result.get("adjudications", [])
    if not isinstance(adjudications, list):
        raise ValueError("stage3 judge output missing adjudications list")
    by_key: dict[tuple[str, int], dict[str, Any]] = {}
    for item in adjudications:
        if not isinstance(item, dict):
            continue
        subtask_id = item.get("subtask_id")
        rubric_index = item.get("rubric_index")
        if isinstance(subtask_id, str) and isinstance(rubric_index, int):
            by_key[(subtask_id, rubric_index)] = item

    adjudicated = 0
    for rubric in result.get("rubric_results", []):
        if not isinstance(rubric, dict) or str(rubric.get("verdict", "")).lower() != "uncertain":
            continue
        key = (str(rubric.get("subtask_id")), int(rubric.get("rubric_index", -1)))
        item = by_key.get(key)
        if item is None:
            rubric["passed"] = False
            rubric["verdict"] = "stage3_fail"
            rubric["stage3"] = {
                "reason": "Stage3 judge did not provide an adjudication for this uncertain rubric.",
                "artifact_evidence": [],
                "confidence": "missing",
            }
            continue
        final_passed = bool(item.get("final_passed"))
        rubric["passed"] = final_passed
        rubric["verdict"] = "stage3_pass" if final_passed else "stage3_fail"
        rubric["stage3"] = {
            "reason": str(item.get("reason", "")),
            "artifact_evidence": [str(x) for x in item.get("artifact_evidence", []) if x is not None]
            if isinstance(item.get("artifact_evidence", []), list)
            else [],
            "confidence": str(item.get("confidence", "unknown")),
        }
        adjudicated += 1

    result["stage3_evaluation"] = {
        "evaluated": True,
        "uncertain_rubrics": sum(
            1 for r in task_eval.get("rubric_results", []) if isinstance(r, dict) and r.get("verdict") == "uncertain"
        ),
        "adjudicated_rubrics": adjudicated,
        "overall_notes": judge_result.get("overall_notes", []),
    }
    recompute_summary(result)
    return result


def main() -> int:
    args = parse_args()
    task_eval_path = Path(args.task_evaluation_json).resolve()
    candidate_dir = Path(args.candidate_dir).resolve()
    output_json = Path(args.output_json).resolve()
    output_dir = output_json.parent

    if not task_eval_path.is_file():
        print(f"task evaluation JSON does not exist: {task_eval_path}", file=sys.stderr)
        return 1
    if not candidate_dir.is_dir():
        print(f"candidate directory does not exist: {candidate_dir}", file=sys.stderr)
        return 1

    task_eval = load_json(task_eval_path)
    uncertain = collect_uncertain_rubrics(task_eval)
    result = deepcopy(task_eval)
    if not uncertain:
        result["stage3_evaluation"] = {"evaluated": False, "uncertain_rubrics": 0, "adjudicated_rubrics": 0}
        write_json(output_json, result)
        return 0

    report_path = Path(str(task_eval.get("report_path", candidate_dir / "report.md")))
    if not report_path.is_file():
        report_path = candidate_dir / "report.md"
    if not report_path.is_file():
        report_path = candidate_dir / "insights.md"
    if not report_path.is_file():
        print(f"report file does not exist for Stage3 under {candidate_dir}", file=sys.stderr)
        return 1

    report_text = read_text_best_effort(report_path, MAX_REPORT_CHARS)
    packet = build_artifact_packet(candidate_dir, uncertain)
    output_dir.mkdir(parents=True, exist_ok=True)
    packet_path = output_dir / "artifact_packet.json"
    prompt_path = output_dir / "stage3_prompt.md"
    raw_output_path = output_dir / "stage3_raw_output.txt"
    api_response_path = output_dir / "stage3_api_response.json"
    write_json(packet_path, packet)

    prompt = build_stage3_prompt(task_eval, uncertain, report_text, packet)
    prompt_path.write_text(prompt, encoding="utf-8")
    raw_output, api_response = run_judge(
        prompt,
        args.judge_api_url,
        args.judge_api_key,
        args.judge_model,
        args.temperature,
        args.max_output_tokens,
    )
    raw_output_path.write_text(raw_output, encoding="utf-8")
    write_json(api_response_path, api_response)

    judge_result = parse_judge_json(raw_output)
    final_result = apply_adjudications(task_eval, judge_result)
    final_result["stage3_evaluation"]["artifact_packet_path"] = str(packet_path)
    final_result["stage3_evaluation"]["judge_prompt_path"] = str(prompt_path)
    final_result["stage3_evaluation"]["judge_raw_output_path"] = str(raw_output_path)
    final_result["stage3_evaluation"]["judge_api_response_path"] = str(api_response_path)
    write_json(output_json, final_result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
